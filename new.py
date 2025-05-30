from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi import Form
import time
import numpy as np
import cv2
import re
from typing import List , Optional
import easyocr
import logging
import asyncio
import datetime
from concurrent.futures import ThreadPoolExecutor
from functools import lru_cache
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from storage import upload_to_mega
from google_sheets_api.sheet import log_driver_stats_to_google_sheets , get_google_sheets_service
from Database.main import log_drivers_stats_to_mongo

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler('app.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

app = FastAPI(
    title="Delivery Statistics Extractor API",
    description="API to extract numerical statistics from delivery app screenshots",
    version="1.0.0"
)

allowed_origins = [
    "http://localhost:3000",  # React development server
    "https://yourfrontenddomain.com",  # Replace with your production frontend domain
]

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],  # Allows all HTTP methods
    allow_headers=["*"],  # Allows all headers
)

# Global variables for models
reader = None
executor = None

# Initialize EasyOCR reader at startup
@app.on_event("startup")
async def startup_event():
    global reader, executor
    try:
        # Initialize EasyOCR with English language - with lower beam width for faster processing
        reader = easyocr.Reader(['en'], gpu=False, beamWidth=3, batch_size=4)
        logger.info("Initialized EasyOCR reader with optimized parameters")
    except Exception as e:
        reader = easyocr.Reader(['en'], gpu=False)
        logger.info(f"Initialized EasyOCR reader without GPU: {str(e)}")
    
    # Initialize thread pool for parallel processing
    executor = ThreadPoolExecutor(max_workers=4)
    logger.info("Initialized thread pool executor")

@app.on_event("shutdown")
async def shutdown_event():
    global executor
    if executor:
        executor.shutdown()
        logger.info("Shut down thread pool executor")

def process_image_version(img_data):
    """Process a single image version with EasyOCR"""
    name, img = img_data
    try:
        # More efficient parameter setting for EasyOCR
        # results = reader.readtext(img, paragraph=False, min_size=10, text_threshold=0.5, link_threshold=0.4, low_text=0.4)
        results = reader.readtext(
            img, 
            paragraph=False,
            min_size=5,  # Lower to catch smaller text
            text_threshold=0.3,  # Lower to detect more text
            link_threshold=0.2,  # Lower for better text grouping
            low_text=0.2,  # Lower to catch low contrast text
            slope_ths=0.1,  # Better handle slightly rotated text
            width_ths=0.5,  # Improved width threshold
            add_margin=0.05  # Add small margin for better character recognition
        )
        text_blocks = []
        for detection in results:
            bbox, text, confidence = detection
            
            # Only include text with reasonable confidence
            if confidence > 0.4 and text.strip():
                # Calculate coordinates and dimensions
                x_min = min(point[0] for point in bbox)
                y_min = min(point[1] for point in bbox)
                x_max = max(point[0] for point in bbox)
                y_max = max(point[1] for point in bbox)
                w = x_max - x_min
                h = y_max - y_min
                
                # Store text block with position and size information
                text_blocks.append({
                    'text': text,
                    'x': int(x_min),
                    'y': int(y_min),
                    'width': int(w),
                    'height': int(h),
                    'area': int(w * h),
                    'conf': confidence,
                    'source': name
                })
        return text_blocks
    except Exception as e:
        logger.warning(f"Error processing {name} image: {str(e)}")
        return []

def preprocess_for_delivery_stats(image):
    # Convert to grayscale
    if len(image.shape) == 3 and image.shape[2] == 3:
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    else:
        gray = image  # Already grayscale

    # Apply adaptive thresholding specifically tuned for this type of image
    thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                                  cv2.THRESH_BINARY, 11, 2)
    
    # Increase contrast for better text recognition
    alpha = 1.5  # Contrast control
    beta = 10    # Brightness control
    enhanced = cv2.convertScaleAbs(gray, alpha=alpha, beta=beta)
    
    return enhanced

def extract_delivery_stats(image_bytes, verbose=False):
    """
    Extract only numerical statistics data from delivery app screenshots using EasyOCR.
    
    Args:
        image_bytes: Image data as bytes
        verbose: Set to True to print debugging information
        
    Returns:
        dict: Extracted numerical statistics with labels
    """
    global reader, executor
    
    try:
        # Decode image
        nparr = np.frombuffer(image_bytes, np.uint8)
        image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        # Apply specialized preprocessing
        image = preprocess_for_delivery_stats(image)
        if image is None:
            raise ValueError("Could not decode image")
        
        # Resize large images for better performance
        max_dimension = 1200
        height, width = image.shape[:2]
        if max(height, width) > max_dimension:
            scale = max_dimension / max(height, width)
            image = cv2.resize(image, None, fx=scale, fy=scale, interpolation=cv2.INTER_AREA)
            height, width = image.shape[:2]
            logger.info(f"Resized image to {width}x{height} for faster processing")
        
        # Process image versions for better OCR results
        processed_images = []
        
        # Original image
        processed_images.append(("original", image))
        
        # Enhanced grayscale version
        if len(image.shape) == 3 and image.shape[2] == 3:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        else:
            gray = image  # Already grayscale
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        enhanced = clahe.apply(gray)
        processed_images.append(("enhanced", enhanced))
        
        # Additional preprocessing for better number recognition
        # Increase contrast for better digit detection
        _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY)
        processed_images.append(("threshold", thresh))
        
        # Extract text from processed images in parallel
        text_blocks = []
        results = list(executor.map(process_image_version, processed_images))
        
        # Combine all text blocks
        for blocks in results:
            text_blocks.extend(blocks)
        
        # Remove duplicates (keep the one with highest confidence)
        unique_blocks = {}
        for block in text_blocks:
            text = block['text'].strip().lower()
            if text not in unique_blocks or block['conf'] > unique_blocks[text]['conf']:
                unique_blocks[text] = block
        
        text_blocks = list(unique_blocks.values())
        
        # Sort blocks by vertical position for better context understanding
        text_blocks.sort(key=lambda x: x['y'])
        
        # Helper function to clean text
        def clean_text(text):
            return re.sub(r'\s+', ' ', text).strip()

        def extract_colon_values(text_blocks):
            stats = {}
            for block in text_blocks:
                text = block['text']
                # Look for patterns like "Label: 123"
                if ':' in text:
                    parts = text.split(':', 1)
                    if len(parts) == 2:
                        label = parts[0].strip()
                        value_text = parts[1].strip()
                        if value_text.isdigit():
                            value = int(value_text)
                            
                            # Normalize the label
                            if "successful deliveries" in label.lower():
                                stats["successful_deliveries"] = value
                            elif "carry forward" in label.lower() and "deliveries" not in label.lower():
                                stats["carry_forward_deliveries"] = value
                            elif "cannot deliver" in label.lower():
                                stats["cannot_deliver"] = value
                            elif "successful collections" in label.lower():
                                stats["successful_collections"] = value
                            elif "cannot collect" in label.lower():
                                stats["cannot_collect"] = value
            return stats    
        
        # Combined text for global pattern matching
        full_text = " ".join([block['text'] for block in text_blocks])
        full_text = clean_text(full_text)
        
        if verbose:
            logger.info(f"Extracted full text: {full_text}")
            for block in text_blocks:
                logger.info(f"Block: {block['text']} at ({block['x']},{block['y']})")
        
        # Define patterns for delivery statistics
        stat_patterns = [
                # Standard patterns
                re.compile(r'(Successful\s*deliveries)[\s:]+(\d+)', re.IGNORECASE),
                re.compile(r'(Carry\s*forward)[\s:]+(\d+)', re.IGNORECASE),
                re.compile(r'(Cannot\s*deliver)[\s:]+(\d+)', re.IGNORECASE),
                re.compile(r'(Successful\s*collections)[\s:]+(\d+)', re.IGNORECASE),
                re.compile(r'(Cannot\s*collect)[\s:]+(\d+)', re.IGNORECASE),
                
                # Fallback patterns with looser matching
                re.compile(r'(Successful\s*deliver).*?(\d+)', re.IGNORECASE),
                re.compile(r'(Carry\s*forward).*?(\d+)', re.IGNORECASE),
                re.compile(r'(Cannot\s*deliver).*?(\d+)', re.IGNORECASE),
                re.compile(r'(Successful\s*collect).*?(\d+)', re.IGNORECASE),
                re.compile(r'(Cannot\s*collect).*?(\d+)', re.IGNORECASE),

                # Exact patterns matching the format in your image
                re.compile(r'Successful\s*deliveries:\s*(\d+)', re.IGNORECASE),
                re.compile(r'Carry\s*forward:\s*(\d+)', re.IGNORECASE),
                re.compile(r'Cannot\s*deliver:\s*(\d+)', re.IGNORECASE),
                re.compile(r'Successful\s*collections:\s*(\d+)', re.IGNORECASE),
                re.compile(r'Cannot\s*collect:\s*(\d+)', re.IGNORECASE),
            ]
        
        # Method 1: Extract stats using regex on full text
        stats = {}
        section = None  # Track whether we're in "deliveries" or "collections" section
        
        for pattern in stat_patterns:
            matches = pattern.finditer(full_text)
            for match in matches:
                # Handle 1-group or 2-group regexes safely
                if match.lastindex == 2:
                    label = clean_text(match.group(1))
                    value = int(match.group(2))
                elif match.lastindex == 1:
                    label = pattern.pattern  # fallback: use the pattern string as a clue
                    value = int(match.group(1))
                else:
                    continue  # skip invalid matches

                # Determine which section this belongs to
                if "deliver" in label.lower():
                    section = "deliveries"
                elif "collect" in label.lower():
                    section = "collections"

                # Normalize the label
                lower_label = label.lower()
                if "successful" in lower_label:
                    if section == "deliveries":
                        stats["successful_deliveries"] = value
                    elif section == "collections":
                        stats["successful_collections"] = value
                elif "carry" in lower_label and "forward" in lower_label:
                    if section == "deliveries":
                        stats["carry_forward_deliveries"] = value
                    elif section == "collections":
                        stats["carry_forward_collections"] = value
                elif "cannot" in lower_label:
                    if "deliver" in lower_label:
                        stats["cannot_deliver"] = value
                    elif "collect" in lower_label:
                        stats["cannot_collect"] = value

        # Add direct colon extraction method 
        colon_stats = extract_colon_values(text_blocks)

        # Add found stats to our overall stats dictionary
        for key, value in colon_stats.items():
            if key not in stats:
                stats[key] = value
        # Method 2: If not all stats found, try analyzing adjacent blocks
        if len(stats) < 5:  # We expect at least 5 statistics
            # Group blocks that are likely part of the same line (similar y-coordinate)
            y_tolerance = 10  # pixels
            lines = []
            current_line = []
            last_y = None
            
            for block in text_blocks:
                if last_y is None or abs(block['y'] - last_y) <= y_tolerance:
                    current_line.append(block)
                else:
                    if current_line:
                        lines.append(current_line)
                    current_line = [block]
                last_y = block['y']
            
            if current_line:
                lines.append(current_line)
            
            # Sort blocks within each line by x-coordinate
            for line in lines:
                line.sort(key=lambda x: x['x'])
            
            # Process each line to find label-value pairs
            for line in lines:
                line_text = " ".join([block['text'] for block in line])
                
                # Try patterns on this line
                for pattern in stat_patterns:
                    match = pattern.search(line_text)
                    if match:
                        label = clean_text(match.group(1))
                        value = int(match.group(2))
                        
                        # Determine which section this belongs to
                        if "deliver" in label.lower() or "deliveries" in line_text.lower():
                            section = "deliveries"
                        elif "collect" in label.lower() or "collections" in line_text.lower():
                            section = "collections"
                        
                        # Normalize the label
                        if "successful" in label.lower():
                            if section == "deliveries":
                                stats["successful_deliveries"] = value
                            elif section == "collections":
                                stats["successful_collections"] = value
                        elif "carry" in label.lower() and "forward" in label.lower():
                            if section == "deliveries":
                                stats["carry_forward_deliveries"] = value
                            elif section == "collections":
                                stats["carry_forward_collections"] = value
                        elif "cannot" in label.lower():
                            if "deliver" in label.lower():
                                stats["cannot_deliver"] = value
                            elif "collect" in label.lower():
                                stats["cannot_collect"] = value
                        break
        
        # Method 3: If still missing stats, use spatial analysis to match labels with numbers
        if len(stats) < 5:
            # Find all blocks containing just numbers
            number_blocks = []
            for block in text_blocks:
                if re.match(r'^\s*\d+\s*$', block['text']):
                    number_blocks.append({
                        'value': int(block['text'].strip()),
                        'x': block['x'],
                        'y': block['y'],
                        'width': block['width'],
                        'height': block['height']
                    })
            
            # Find all potential label blocks
            label_keywords = {
                'successful_deliveries': ['successful', 'deliveries'],
                'carry_forward_deliveries': ['carry', 'forward', 'deliveries'],
                'cannot_deliver': ['cannot', 'deliver'],
                'successful_collections': ['successful', 'collections'],
                'carry_forward_collections': ['carry', 'forward', 'collections'],
                'cannot_collect': ['cannot', 'collect']
            }
            
            # Try to match labels with nearby numbers
            for key, keywords in label_keywords.items():
                if key in stats:
                    continue  # Already found this stat
                
                # Find blocks containing these keywords
                for block in text_blocks:
                    block_text = block['text'].lower()
                    if any(keyword in block_text for keyword in keywords):
                        # Look for the closest number block
                        closest_num = None
                        min_dist = float('inf')
                        
                        for num_block in number_blocks:
                            # Calculate distance (emphasizing horizontal alignment)
                            h_dist = abs(num_block['x'] - (block['x'] + block['width']))
                            v_dist = abs(num_block['y'] - block['y'])
                            
                            # Prioritize numbers to the right of labels and on same line
                            if num_block['x'] > block['x'] and v_dist < 30:
                                dist = h_dist + (v_dist * 3)
                                if dist < min_dist:
                                    min_dist = dist
                                    closest_num = num_block
                        
                        if closest_num and min_dist < 200:
                            stats[key] = closest_num['value']
                            break
        
        # Format the results with human-readable labels
        delivery_stats = {
            "Successful deliveries": stats.get("successful_deliveries"),
            "Carry forward deliveries": stats.get("carry_forward_deliveries"),
            "Cannot deliver": stats.get("cannot_deliver"),
            "Successful collections": stats.get("successful_collections"),
            "Carry forward collections": stats.get("carry_forward_collections"),
            "Cannot collect": stats.get("cannot_collect")
        }
        
        # Remove None values
        delivery_stats = {k: v for k, v in delivery_stats.items() if v is not None}
        
        return delivery_stats
        
    except Exception as e:
        logger.error(f"Error processing image: {str(e)}")
        raise ValueError(f"Image processing failed: {str(e)}")

def log_driver_stats_to_excel(
    filename: str,
    date: str,
    driver_name: str,
    warehouse: str,
    route: str,
    successful_deliveries: int,
    successful_collections: int,
    image_link: str
 ):

    # Column headers for each driver's block
    stat_headers = [
        "Successful Deliveries", "Successful Collections",
        "Total Jobs", "Route", "Image Link"
    ]

    # Create or load workbook
    if not os.path.exists(filename):
        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet
    else:
        wb = load_workbook(filename)

    # Get or create sheet by warehouse name
    if warehouse not in wb.sheetnames:
        ws = wb.create_sheet(title=warehouse)
        # Create initial header row with 'Date' in A1 and A2
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.cell(row=1, column=1, value="Date").font = Font(bold=True)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    else:
        ws = wb[warehouse]

    # Map driver name to column start position
    driver_columns = {}
    col = 2  # Starting from column B (since A is Date)
    while ws.cell(row=1, column=col).value:
        driver = ws.cell(row=1, column=col).value
        if driver:
            driver_columns[driver] = col
            col += len(stat_headers)
        else:
            break

    # If driver not already in headers, add them
    if driver_name not in driver_columns:
        start_col = col
        driver_columns[driver_name] = start_col
        end_col = start_col + len(stat_headers) - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        ws.cell(row=1, column=start_col, value=driver_name).font = Font(bold=True)
        ws.cell(row=1, column=start_col).alignment = Alignment(horizontal='center')

        for i, header in enumerate(stat_headers):
            ws.cell(row=2, column=start_col + i, value=header).font = Font(bold=True)
            ws.cell(row=2, column=start_col + i).alignment = Alignment(horizontal='center')

    # 👉 Append new row even if the date already exists (to support multiple routes per driver per date)
    new_row = ws.max_row + 1

    # Write date in column A
    ws.cell(row=new_row, column=1, value=date)
    ws.cell(row=new_row, column=1).alignment = Alignment(horizontal='center')

    # Insert stats
    start_col = driver_columns[driver_name]
    total_jobs = successful_deliveries + successful_collections
    values = [successful_deliveries, successful_collections, total_jobs, route, image_link]

    for i, val in enumerate(values):
        cell = ws.cell(row=new_row, column=start_col + i, value=val)
        cell.alignment = Alignment(horizontal='center')
        if i == 4:  # Image Link column
            cell.hyperlink = image_link
            cell.font = Font(color="0000FF", underline="single")

    # Auto-adjust column widths
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(filename)
    print(f"Appended data for driver '{driver_name}' on {date} to warehouse sheet '{warehouse}'.")

@app.post("/extract-stats")
async def extract_stats(files: List[UploadFile] = File(...)):
    """
    Extract numerical statistics from delivery app screenshots
    
    Args:
        files: The uploaded image files
        
    Returns:
        dict: Extracted statistics with processing time
    """
    try:
        results = []
        for file in files:
            # Record start time for performance measurement
            start_time = time.time()
            
            # Check file type
            content_type = file.content_type
            if not content_type.startswith("image/"):
                raise HTTPException(status_code=400, detail="File must be an image")
            
            # Read the image bytes
            image_bytes = await file.read()
            
            # Process the image to extract numerical data only
            result = extract_delivery_stats(image_bytes)
            
            # Calculate processing time
            end_time = time.time()
            result["processing_time"] = f"{end_time - start_time:.4f} seconds"
            
            results.append(result)
        
        return {"extracted_statistics": results}
    
    except ValueError as e:
        logger.error(f"Value error: {str(e)}")
        raise HTTPException(status_code=422, detail=str(e))
    
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        raise HTTPException(status_code=500, detail="An unexpected error occurred during processing")

# @app.post("/extract-excel")
# async def extract_stats(
#     files: List[UploadFile] = File(...),
#     driver_name: str = Form(...),
#     warehouse: str = Form(...),
#     routes: List[str] = Form(...),  
#     date: str = Form(...),
#     excel_filename: Optional[str] = Form("driver_stats.xlsx")
#  ):
#     """
#     Extract numerical statistics and log them to Excel per driver and route.
#     """
#     try:
#         results = []
#         if len(files) > len(routes):
#             raise ValueError("Cannot have more images than routes")

#         for idx, file in enumerate(files):
#             start_time = time.time()

#             if not file.content_type.startswith("image/"):
#                 raise HTTPException(status_code=400, detail=f"File {file.filename} must be an image")

#             image_bytes = await file.read()
#             stats = extract_delivery_stats(image_bytes)

#             end_time = time.time()
#             stats["processing_time"] = f"{end_time - start_time:.4f} seconds"

#             public_image_url = upload_to_mega(file.filename , image_bytes)

#             # Log to Excel
#             excel_file = excel_filename or "driver_stats1.xlsx"
#             if len(files) == 1 and len(routes) > 1:
#                 for route in routes:
#                     try:
#                         success = log_driver_stats_to_google_sheets(
#                             date=date,
#                             driver_name=driver_name,
#                             warehouse=warehouse,
#                             route=route,
#                             successful_deliveries=stats.get("Successful deliveries", 0),
#                             successful_collections=stats.get("Successful collections", 0),
#                             image_link = public_image_url
#                         )
#                         stats["sheets_status"] = "Success" if success else "Failed"
#                     except ValueError as ve:
#                         stats["sheets_status"] = str(ve)
#             else:
#                 try:
#                     success = log_driver_stats_to_google_sheets(
#                         date=date,
#                         driver_name=driver_name,
#                         warehouse=warehouse,
#                         route=routes[idx],
#                         successful_deliveries=stats.get("Successful deliveries", 0),
#                         successful_collections=stats.get("Successful collections", 0),
#                         image_link = public_image_url
#                     )
#                     stats["sheets_status"] = "Success" if success else "Failed"
#                 except ValueError as ve:
#                     stats["sheets_status"] = str(ve)
                
#             stats["image_link"] = public_image_url
#             results.append(stats)

#         return {"extracted_statistics": results}

#     except ValueError as e:
#         logger.error(f"Value error: {str(e)}")
#         raise HTTPException(status_code=422, detail=str(e))
#     except Exception as e:
#         logger.error(f"Unexpected error: {str(e)}")
#         raise HTTPException(status_code=500, detail="Unexpected error during processing")

@app.post("/extract-excel")
async def extract_stats(
    files: List[UploadFile] = File(...),
    driver_name: str = Form(...),
    warehouse: str = Form(...),
    routes: List[str] = Form(...),  
    date: str = Form(...)
 ):
    """
    Extract stats and log them to Google Sheets with improved error handling
    """
    try:
        # Validate inputs first
        if not files:
            raise HTTPException(status_code=400, detail="No files provided")
        if not routes:
            raise HTTPException(status_code=400, detail="No routes provided")
        if len(files) > len(routes):
            raise HTTPException(status_code=400, detail="Cannot have more images than routes")

        # Initialize Google Sheets service once
        try:
            sheets_service = get_google_sheets_service()
            logger.info("Successfully initialized Google Sheets service")
        except Exception as e:
            logger.error(f"Failed to initialize Google Sheets service: {str(e)}")
            raise HTTPException(
                status_code=500, 
                detail="Failed to connect to Google Sheets. Please try again later."
            )

        results = []
        
        for idx, file in enumerate(files):
            try:
                # Validate file type
                if not file.content_type.startswith("image/"):
                    raise HTTPException(
                        status_code=400, 
                        detail=f"File {file.filename} must be an image"
                    )

                # Process image with timing
                start_time = time.time()
                image_bytes = await file.read()
                
                try:
                    stats = extract_delivery_stats(image_bytes)
                except Exception as e:
                    logger.error(f"Error extracting stats from image: {str(e)}")
                    raise HTTPException(
                        status_code=422,
                        detail=f"Failed to extract stats from image {file.filename}: {str(e)}"
                    )

                processing_time = time.time() - start_time
                stats["processing_time"] = f"{processing_time:.4f} seconds"

                # Upload image with retry mechanism
                try:
                    public_image_url = upload_to_mega(file.filename, image_bytes)
                except Exception as e:
                    logger.error(f"Failed to upload image to storage: {str(e)}")
                    raise HTTPException(
                        status_code=500,
                        detail=f"Failed to upload image {file.filename} to storage"
                    )

                # Handle Google Sheets updates with retries
                max_retries = 3
                retry_delay = 2  # seconds
                
                if len(files) == 1 and len(routes) > 1:
                    # Single image, multiple routes case
                    for route in routes:
                        for attempt in range(max_retries):
                            try:
                                success = log_driver_stats_to_google_sheets(
                                    date=date,
                                    driver_name=driver_name,
                                    warehouse=warehouse,
                                    route=route,
                                    successful_deliveries=stats.get("Successful deliveries", 0),
                                    successful_collections=stats.get("Successful collections", 0),
                                    image_link=public_image_url
                                )
                                if success:
                                    stats["sheets_status"] = "Success"
                                    # Only try MongoDB if Sheets was successful
                                    mongo_success = log_drivers_stats_to_mongo(
                                        date=date,
                                        driver_name=driver_name,
                                        warehouse=warehouse,
                                        route=route,
                                        successful_deliveries=stats.get("Successful deliveries", 0),
                                        successful_collections=stats.get("Successful collections", 0),
                                        image_link=public_image_url
                                    )
                                    
                                    stats["mongo_status"] = "Success" if mongo_success else "Failed"
                                    if mongo_success:
                                        stats["mongo_status"] = "Success"
                                        break
                                else:
                                    stats["sheets_status"] = "Failed"
                            except Exception as e:
                                logger.warning(f"Attempt {attempt + 1} failed: {str(e)}")
                                if attempt == max_retries - 1:
                                    stats["sheets_status"] = f"Failed after {max_retries} attempts: {str(e)}"
                                else:
                                    await asyncio.sleep(retry_delay)
                else:
                    # Multiple images, one route each case
                    for attempt in range(max_retries):
                        try:
                            success = log_driver_stats_to_google_sheets(
                                date=date,
                                driver_name=driver_name,
                                warehouse=warehouse,
                                route=routes[idx],
                                successful_deliveries=stats.get("Successful deliveries", 0),
                                successful_collections=stats.get("Successful collections", 0),
                                image_link=public_image_url
                            )
                            if success:
                                stats["sheets_status"] = "Success"
                                # Only try MongoDB if Sheets was successful
                                mongo_success = log_drivers_stats_to_mongo(
                                    date=date,
                                    driver_name=driver_name,
                                    warehouse=warehouse,
                                    route=routes[idx],
                                    successful_deliveries=stats.get("Successful deliveries", 0),
                                    successful_collections=stats.get("Successful collections", 0),
                                    image_link=public_image_url
                                )
                                
                                stats["mongo_status"] = "Success" if mongo_success else "Failed"
                                if mongo_success:
                                    stats["mongo_status"] = "Success"
                                    break
                            else:
                                stats["sheets_status"] = "Failed"
                                stats["mongo_status"] = "Skipped due to sheet failure"
                        except Exception as e:
                            logger.warning(f"Attempt {attempt + 1} failed: {str(e)}")
                            if attempt == max_retries - 1:
                                stats["sheets_status"] = f"Failed after {max_retries} attempts: {str(e)}"
                                stats["mongo_status"] = "Skipped due to sheet failure"
                            else:
                                await asyncio.sleep(retry_delay)

                stats["image_link"] = public_image_url
                results.append(stats)

            except Exception as file_error:
                # Handle individual file processing errors
                logger.error(f"Error processing file {file.filename}: {str(file_error)}")
                results.append({
                    "filename": file.filename,
                    "error": str(file_error),
                    "status": "Failed"
                })

        return {
            "extracted_statistics": results,
            "total_files_processed": len(files),
            "successful_processes": len([r for r in results if r.get("sheets_status") == "Success"]),
            "timestamp": datetime.datetime.now().isoformat()
        }

    except HTTPException as http_error:
        # Re-raise HTTP exceptions
        raise http_error
    except Exception as e:
        # Log unexpected errors and return a generic error message
        logger.error(f"Unexpected error in extract_stats: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail="An unexpected error occurred. Please try again later."
        )

# Health check endpoint
@app.get("/health")
async def health_check():
    """Simple health check endpoint"""
    return {"status": "ok", "message": "Delivery Stats Extractor API is running"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)